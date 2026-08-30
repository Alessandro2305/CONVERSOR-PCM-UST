import io
import re
import base64
import traceback
import openpyxl
from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pypdf import PdfReader, PdfWriter
from reportlab.pdfgen import canvas
from reportlab.lib import colors

app = FastAPI(redirect_slashes=False)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

def extrair_apenas_codigo_limpo(texto) -> str:
    """ Remove CNH, CASE, espaços e caracteres especiais """
    if not texto:
        return ""
    txt = str(texto).upper().strip()
    txt = re.sub(r'CNH$', '', txt)
    txt = re.sub(r'^CASE[-_\s]*', '', txt)
    return re.sub(r'[^A-Z0-9]', '', txt)

@app.get("/")
@app.get("/api")
def root():
    return {"status": "ok", "message": "API Conversor PCM UST Operacional"}

@app.post("/escrever-no-pdf-original/")
@app.post("/escrever-no-pdf-original")
@app.post("/api/escrever-no-pdf-original/")
@app.post("/api/escrever-no-pdf-original")
async def escrever_no_pdf_original(
    pdf_file: UploadFile = File(...),
    excel_depara: UploadFile = File(...)
):
    try:
        # 1. Leitura Completa do Excel
        excel_bytes = await excel_depara.read()
        wb = openpyxl.load_workbook(filename=io.BytesIO(excel_bytes), data_only=True)

        mapa_sol = {}
        mapa_desc = {}

        sheet = wb['C'] if 'C' in wb.sheetnames else wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or all(v is None for v in row):
                continue

            raw_sol = str(row[0] or '').replace(".0", "").strip()
            raw_desc = str(row[1] or '').strip() if len(row) > 1 and row[1] else "SEM DESCRIÇÃO"

            if not raw_sol or raw_sol.lower() in ["none", "nan"]:
                continue

            indices = [2, 1, 0] + list(range(3, len(row)))
            for idx in indices:
                if idx < len(row) and row[idx] is not None:
                    val_str = str(row[idx])
                    chave_limpa = extrair_apenas_codigo_limpo(val_str)
                    
                    if chave_limpa and len(chave_limpa) >= 3 and chave_limpa not in ["NONE", "NAN"]:
                        if chave_limpa not in mapa_sol:
                            mapa_sol[chave_limpa] = raw_sol
                            mapa_desc[chave_limpa] = raw_desc

        # 2. Leitura e Inserção Precisa no PDF
        pdf_bytes = await pdf_file.read()
        reader_base = PdfReader(io.BytesIO(pdf_bytes))
        writer = PdfWriter()

        itens_encontrados = []
        codigos_processados = set()

        for page in reader_base.pages:
            try:
                page_width = float(page.mediabox.width)
                page_height = float(page.mediabox.height)

                # Mapeia dinamicamente o Y da linha do título "PEÇAS, COMPONENTES"
                y_inicio_tabela = 0

                def MapearCabecalho(text, cm, tm, font_dict, font_size):
                    nonlocal y_inicio_tabela
                    if not text:
                        return
                    t = str(text).upper()
                    if "PECAS" in t or "LUBRIFICANTES" in t or "CÓDIGO" in t or "CODIGO" in t:
                        matrix = tm if tm is not None and len(tm) >= 6 else cm
                        if matrix and len(matrix) >= 6:
                            y = matrix[5]
                            if y_inicio_tabela == 0 or y < y_inicio_tabela:
                                y_inicio_tabela = y

                # Passagem 1: Encontra a altura Y exata de onde começa a tabela
                page.extract_text(visitor_text=MapearCabecalho)

                # Se não achou a palavra-chave, usa como piso padrão os 45% inferiores da folha
                if y_inicio_tabela == 0:
                    y_inicio_tabela = page_height * 0.55

                packet = io.BytesIO()
                can = canvas.Canvas(packet, pagesize=(page_width, page_height))
                escreveu_algo = False

                blocos_para_escrever = []

                def visitor_body(text, cm, tm, font_dict, font_size):
                    if not text:
                        return
                    t_bruto = str(text).strip()
                    if not t_bruto:
                        return

                    matrix = tm if tm is not None and len(matrix) >= 6 else cm
                    x_pos = matrix[4] if matrix and len(matrix) >= 6 else 0
                    y_pos = matrix[5] if matrix and len(matrix) >= 6 else 0

                    # REGRA RIGOROSA DE TRAVA:
                    # 1. Posição X restrita estritamente à coluna de código (X entre 35 e 95)
                    # 2. Posição Y estritamente ABAIXO da linha de início da tabela (y_pos < y_inicio_tabela - 15)
                    if 35 <= x_pos <= 95 and y_pos < (y_inicio_tabela - 15):
                        
                        cod_limpo = extrair_apenas_codigo_limpo(t_bruto)

                        # Evita capturar cabeçalhos de coluna
                        if "CODIGO" in cod_limpo or "PECAS" in cod_limpo or len(cod_limpo) < 4:
                            return

                        if cod_limpo in mapa_sol:
                            raw_sol = mapa_sol[cod_limpo]
                            descricao = mapa_desc.get(cod_limpo, "SEM DESCRIÇÃO")
                            cod_sol = f"SOL-{raw_sol}" if not raw_sol.startswith("SOL") else raw_sol

                            if cod_limpo not in codigos_processados:
                                codigos_processados.add(cod_limpo)
                                itens_encontrados.append({
                                    "status": "Convertido",
                                    "codigo_original": t_bruto,
                                    "codigo_sol": cod_sol,
                                    "descricao": descricao
                                })

                            # Posição X = 92 desenha o SOL exatamente ao lado direito do código CNH
                            blocos_para_escrever.append((92, y_pos, cod_sol))

                        elif cod_limpo not in codigos_processados and not t_bruto.isdigit():
                            codigos_processados.add(cod_limpo)
                            itens_encontrados.append({
                                "status": "Não encontrado",
                                "codigo_original": t_bruto,
                                "codigo_sol": "—",
                                "descricao": "SEM DESCRIÇÃO"
                            })

                page.extract_text(visitor_text=visitor_body)

                # Escreve com tarja branca de fundo no texto para evitar sobreposição
                if blocos_para_escrever:
                    for x, y, texto_sol in blocos_para_escrever:
                        # Fundo limpo
                        can.setFillColor(colors.white)
                        can.rect(x - 2, y - 1, 55, 8, fill=1, stroke=0)
                        
                        # Texto em Azul
                        can.setFont("Helvetica-Bold", 6.5)
                        can.setFillColor(colors.HexColor("#0284c7"))
                        can.drawString(x, y, texto_sol)
                        escreveu_algo = True

                if escreveu_algo:
                    can.save()
                    packet.seek(0)
                    overlay_pdf = PdfReader(packet)
                    if len(overlay_pdf.pages) > 0:
                        page.merge_page(overlay_pdf.pages[0])

                writer.add_page(page)

            except Exception:
                writer.add_page(page)

        output_stream = io.BytesIO()
        writer.write(output_stream)
        output_stream.seek(0)

        pdf_b64 = base64.b64encode(output_stream.getvalue()).decode('utf-8')

        return {
            "pdf_base64": pdf_b64,
            "itens": itens_encontrados
        }

    except HTTPException as http_exc:
        raise http_exc
    except Exception as e:
        tb = traceback.format_exc()
        print(f"CRITICAL ERROR: {tb}")
        raise HTTPException(status_code=500, detail=f"Erro interno no servidor: {str(e)}")